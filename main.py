import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    
    wb = xl.load_workbook('transaction.xlsx') # Stores the target workbook inside a variable 'wb'
    sheet = wb['Sheet1'] # Acesses the specific sheet that contains the targeted data

    for row in range(2, sheet.max_row + 1): # Loops through rows 2 to 4 inclusive
        cell = sheet.cell(row, 3) # Stores the data of the cell at [row, 3] inside a variable 'cell'
        corrected_price = cell.value * 0.9 # Stores the value of the current cell after a 10% discount is applied in a variable 'corrected_price'
        corrected_price_cell = sheet.cell(row, 4) # Sets a variable 'corrected_price_cell' as a cell in a new column to store the corrected prices
        corrected_price_cell.value = corrected_price # Sets the value of 'corrected_price_cell' to the corrected_price variable

    values = Reference(sheet, min_row= 2, max_row= sheet.max_row, min_col= 4, max_col= 4) # Stores the values of the 4th column inside a variable 'values'

    chart = BarChart() # Creates an instance of the BarChart class
    chart.add_data(values) # Feeds the values from 'values' to the chart
    sheet.add_chart(chart, 'e2') # Adds the 'chart' variable to the spreadsheet with the upper left corner at 'e2'

    wb.save('transactions2.xlsx') # Saves the changes to a new spreadsheet to avoid overriding the original file as a failsafe in case bugs exist
